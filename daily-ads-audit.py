#!/usr/bin/env python3
"""
Daily Google Ads + GA4 Audit — 3 Clients
Generates XLSX, PDF, HTML per client
Output: ~/Downloads/[Client]/[YYYY-MM-DD]/[HH-MM-SS]/
"""
import os, sys, json, requests, datetime, math
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ─── CONFIG ─────────────────────────────────────────────────────────────────
WINDSOR_API_KEY = os.environ.get("WINDSOR_API_KEY", "YOUR_WINDSOR_API_KEY_HERE")
WINDSOR_BASE    = "https://connectors.windsor.ai/all"
DOWNLOADS       = Path.home() / "Downloads"

CLIENTS = {
    "Ortho": {
        "full_name": "Orthopedic Surgery Practice",
        "color":     "#1a4f8a",
        "accent":    "#e8f0fb",
        "gads_accounts": ["106-254-8978"],
        "ga4_accounts":  ["258012138", "439336281"],
        "ga4_names":     {"258012138": "cityorthosports.com", "439336281": "newjerseykneesurgeon"},
    },
    "HeatWeave": {
        "full_name": "HeatWeave HVAC",
        "color":     "#c0392b",
        "accent":    "#fdf2f1",
        "gads_accounts": ["494-668-8111"],
        "ga4_accounts":  ["375400573"],
        "ga4_names":     {"375400573": "Heatwave Florida"},
    },
    "Tack": {
        "full_name": "Tack Media",
        "color":     "#1e8449",
        "accent":    "#eafaf1",
        "gads_accounts": [],   # not connected — manual note
        "ga4_accounts":  [],
        "ga4_names":     {},
    },
}

TODAY      = datetime.date.today()
DATE_FROM  = (TODAY - datetime.timedelta(days=30)).isoformat()
DATE_TO    = TODAY.isoformat()
DATE_PREV  = (TODAY - datetime.timedelta(days=60)).isoformat()
DATE_PREV_TO = (TODAY - datetime.timedelta(days=31)).isoformat()

# ─── WINDSOR DATA PULL ───────────────────────────────────────────────────────
GADS_FIELDS = [
    "date","account_id","account_name",
    "campaign","campaign_status","campaign_bidding_strategy_type",
    "ad_group","ad_group_status",
    "spend","clicks","impressions","conversions","conversion_value",
    "cpc","ctr","cpm","roas",
]

GA4_FIELDS = [
    "date","account_id",
    "sessions","active_users","conversions","bounce_rate",
    "average_session_duration","engaged_sessions",
    "default_channel_group","event_name",
]


def windsor_pull(connector, fields, accounts, date_from, date_to):
    if not accounts:
        return []
    params = {
        "api_key":   WINDSOR_API_KEY,
        "connector": connector,
        "date_from": date_from,
        "date_to":   date_to,
        "fields":    ",".join(fields),
        "account_id": ",".join(accounts),
    }
    try:
        r = requests.get(WINDSOR_BASE, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        return data.get("data", data) if isinstance(data, dict) else data
    except Exception as e:
        print(f"  Windsor error ({connector}): {e}")
        return []


def safe(val, default=0):
    try:
        return float(val) if val not in (None, "", "N/A") else default
    except:
        return default


def pct_change(curr, prev):
    if prev == 0:
        return 0
    return ((curr - prev) / abs(prev)) * 100


# ─── ANALYSIS ENGINE ─────────────────────────────────────────────────────────
def analyze_gads(rows):
    """Returns summary + issues + recommendations."""
    if not rows:
        return {"error": "No Google Ads data — account not connected or no spend"}

    total_spend      = sum(safe(r.get("spend")) for r in rows)
    total_clicks     = sum(safe(r.get("clicks")) for r in rows)
    total_impr       = sum(safe(r.get("impressions")) for r in rows)
    total_conv       = sum(safe(r.get("conversions")) for r in rows)
    total_conv_val   = sum(safe(r.get("conversion_value")) for r in rows)

    avg_cpc  = total_spend / total_clicks   if total_clicks  else 0
    avg_ctr  = total_clicks / total_impr    if total_impr    else 0
    avg_cpa  = total_spend / total_conv     if total_conv    else 0
    avg_roas = total_conv_val / total_spend if total_spend   else 0

    # Campaign breakdown
    camp_map = {}
    for r in rows:
        c = r.get("campaign", "Unknown")
        if c not in camp_map:
            camp_map[c] = {"spend":0,"clicks":0,"impressions":0,"conversions":0,"status":r.get("campaign_status",""),"strategy":r.get("campaign_bidding_strategy_type","")}
        camp_map[c]["spend"]       += safe(r.get("spend"))
        camp_map[c]["clicks"]      += safe(r.get("clicks"))
        camp_map[c]["impressions"] += safe(r.get("impressions"))
        camp_map[c]["conversions"] += safe(r.get("conversions"))

    campaigns = []
    for name, d in camp_map.items():
        d["name"] = name
        d["cpc"]  = d["spend"] / d["clicks"]      if d["clicks"]      else 0
        d["ctr"]  = d["clicks"] / d["impressions"] if d["impressions"] else 0
        d["cpa"]  = d["spend"] / d["conversions"]  if d["conversions"] else 0
        campaigns.append(d)
    campaigns.sort(key=lambda x: x["spend"], reverse=True)

    # Client-specific benchmark overrides
    # Ortho: CPA <$80, CTR 5%+, CPC $4-15, QS 7+, IS 40%+ (CityOrtho_GoogleAds_Workbook_v2.xlsx)
    client_benchmarks = rows[0].get("__client_benchmarks__", {}) if rows else {}
    cpa_target  = client_benchmarks.get("cpa_target",  80)
    ctr_target  = client_benchmarks.get("ctr_target",  0.05)
    is_target   = client_benchmarks.get("is_target",   0.40)

    # Issue detection
    issues = []
    for c in campaigns:
        if c["impressions"] > 0 and c["ctr"] < 0.02:
            issues.append({
                "priority": "HIGH",
                "issue": f"Low CTR on '{c['name']}'",
                "detail": f"CTR = {c['ctr']:.2%} (benchmark: 2-5%)",
                "fix": "Rewrite ad headlines with stronger CTAs, add price extensions, test RSA pinning"
            })
        if c["clicks"] > 50 and c["conversions"] == 0:
            issues.append({
                "priority": "HIGH",
                "issue": f"Zero conversions on '{c['name']}'",
                "detail": f"${c['spend']:.0f} spent, {int(c['clicks'])} clicks, 0 conversions",
                "fix": "Check conversion tracking, review landing page alignment, pause or restructure"
            })
        if c["cpa"] > 0 and c["cpa"] > avg_cpa * 2.5:
            issues.append({
                "priority": "MEDIUM",
                "issue": f"High CPA on '{c['name']}'",
                "detail": f"CPA ${c['cpa']:.2f} vs account avg ${avg_cpa:.2f}",
                "fix": "Lower bids, tighten targeting, improve Quality Score, prune poor-performing ad groups"
            })
        if c["cpc"] > avg_cpc * 2:
            issues.append({
                "priority": "MEDIUM",
                "issue": f"Elevated CPC on '{c['name']}'",
                "detail": f"CPC ${c['cpc']:.2f} vs account avg ${avg_cpc:.2f}",
                "fix": "Add negatives, improve relevance score, use Target CPA instead of manual bidding"
            })

    # Ortho-specific: CPA target $80 (from KPI workbook)
    if avg_cpa > cpa_target and total_conv > 0:
        issues.append({
            "priority": "HIGH",
            "issue": f"CPA above client target (${cpa_target})",
            "detail": f"Current CPA ${avg_cpa:.2f} vs target <${cpa_target} (City Ortho benchmark)",
            "fix": "Pause low-conv ad groups, tighten geo (East Rutherford/West Orange/Paramus), improve landing page conversion rate"
        })
    # Ortho-specific: CTR target 5% (healthcare benchmark)
    if total_impr > 500 and avg_ctr < ctr_target:
        issues.append({
            "priority": "MEDIUM",
            "issue": f"CTR below healthcare benchmark ({ctr_target:.0%})",
            "detail": f"Account CTR {avg_ctr:.2%} — NJ ortho benchmark is 5%+",
            "fix": "Strengthen RSA headlines (injury type + location + board-certified), add callout extensions, test ad scheduling"
        })
    # Regenerative Medicine zero conversions (expensive niche)
    for c in campaigns:
        if "regenerat" in c["name"].lower() and c["spend"] > 50 and c["conversions"] == 0:
            issues.append({
                "priority": "CRITICAL",
                "issue": "Regenerative Medicine campaign: spend with zero conversions",
                "detail": f"${c['spend']:.0f} spent on PRP/stem cell keywords — 0 booked appointments",
                "fix": "Review landing page (dedicated PRP page?), check Gravity Forms conversion tag, reduce bids on broad stem cell terms"
            })
    # HIPAA: check if call conversions tracked
    call_conv = sum(safe(r.get("conversions")) for r in rows if "call" in str(r.get("conversion_action_name","")).lower())
    if call_conv == 0 and total_spend > 200:
        issues.append({
            "priority": "HIGH",
            "issue": "No call conversions tracked",
            "detail": "Phone calls (201-500-9450 / 201-613-3388) not appearing as conversions — tracking gap",
            "fix": "Verify Google Ads call extension + call conversion tag in GTM-K96NFSF, check call tracking in GA4"
        })

    if avg_roas > 0 and avg_roas < 2:
        issues.append({
            "priority": "HIGH",
            "issue": "Account-wide ROAS below 2x",
            "detail": f"ROAS = {avg_roas:.2f}x — not profitable for most industries",
            "fix": "Shift budget to top-converting campaigns, pause low ROAS campaigns, raise revenue-per-conversion"
        })
    if total_conv == 0 and total_spend > 100:
        issues.append({
            "priority": "CRITICAL",
            "issue": "Zero conversions account-wide",
            "detail": f"${total_spend:.0f} spent with no recorded conversions",
            "fix": "URGENT: Verify Google Ads conversion tracking via Tag Assistant, check GA4 goal linkage"
        })

    return {
        "total_spend":    total_spend,
        "total_clicks":   total_clicks,
        "total_impr":     total_impr,
        "total_conv":     total_conv,
        "total_conv_val": total_conv_val,
        "avg_cpc":        avg_cpc,
        "avg_ctr":        avg_ctr,
        "avg_cpa":        avg_cpa,
        "avg_roas":       avg_roas,
        "campaigns":      campaigns,
        "issues":         sorted(issues, key=lambda x: {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3}[x["priority"]]),
    }


def analyze_ga4(rows):
    if not rows:
        return {"error": "No GA4 data — property not connected"}

    total_sessions = sum(safe(r.get("sessions")) for r in rows)
    total_users    = sum(safe(r.get("active_users")) for r in rows)
    total_conv     = sum(safe(r.get("conversions")) for r in rows)
    avg_bounce     = sum(safe(r.get("bounce_rate")) for r in rows) / len(rows) if rows else 0
    avg_dur        = sum(safe(r.get("average_session_duration")) for r in rows) / len(rows) if rows else 0

    channels = {}
    for r in rows:
        ch = r.get("default_channel_group", "Unknown")
        if ch not in channels:
            channels[ch] = {"sessions":0,"conversions":0}
        channels[ch]["sessions"]    += safe(r.get("sessions"))
        channels[ch]["conversions"] += safe(r.get("conversions"))

    issues = []
    if avg_bounce > 0.7:
        issues.append({"priority":"HIGH","issue":"High bounce rate","detail":f"{avg_bounce:.1%} bounce rate (benchmark <60%)","fix":"Improve landing page speed, above-fold clarity, mobile UX"})
    if avg_dur < 60:
        issues.append({"priority":"MEDIUM","issue":"Low session duration","detail":f"Avg {avg_dur:.0f}s (benchmark >90s)","fix":"Improve content depth, internal linking, reduce pop-up interruptions"})
    if total_conv == 0:
        issues.append({"priority":"CRITICAL","issue":"Zero GA4 conversions","detail":"No events configured as conversions","fix":"Mark key_event in GA4 admin, set up lead_form, phone_click, purchase events"})

    return {
        "total_sessions": total_sessions,
        "total_users":    total_users,
        "total_conv":     total_conv,
        "avg_bounce":     avg_bounce,
        "avg_dur":        avg_dur,
        "channels":       channels,
        "issues":         issues,
    }


# ─── XLSX BUILDER ────────────────────────────────────────────────────────────
def build_xlsx(client_key, client_cfg, gads, ga4, out_dir):
    wb = openpyxl.Workbook()

    def header_style(ws, row, cols, title, color):
        ws.row_dimensions[row].height = 28
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=color.replace("#",""))
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).value = title

    def write_kpi_row(ws, r, label, val, fmt="{:.0f}", trend=""):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=fmt.format(val) if isinstance(val, (int, float)) else val)
        ws.cell(row=r, column=3, value=trend)

    color = client_cfg["color"].replace("#","")

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    header_style(ws, 1, 3, f"{client_cfg['full_name']} — 30-Day Audit Summary ({TODAY})", color)
    row = 3

    if "error" not in gads:
        ws.cell(row=row, column=1, value="─── GOOGLE ADS ───").font = Font(bold=True, size=11)
        row += 1
        kpis = [
            ("Total Spend",       gads["total_spend"],   "${:.2f}"),
            ("Total Clicks",      gads["total_clicks"],  "{:,.0f}"),
            ("Impressions",       gads["total_impr"],    "{:,.0f}"),
            ("Conversions",       gads["total_conv"],    "{:.1f}"),
            ("Avg CPC",           gads["avg_cpc"],       "${:.2f}"),
            ("Avg CTR",           gads["avg_ctr"],       "{:.2%}"),
            ("Avg CPA",           gads["avg_cpa"],       "${:.2f}"),
            ("ROAS",              gads["avg_roas"],       "{:.2f}x"),
        ]
        for label, val, fmt in kpis:
            write_kpi_row(ws, row, label, val, fmt)
            row += 1
        row += 1

    if "error" not in ga4:
        ws.cell(row=row, column=1, value="─── GA4 ───").font = Font(bold=True, size=11)
        row += 1
        g4kpis = [
            ("Sessions",          ga4["total_sessions"], "{:,.0f}"),
            ("Active Users",      ga4["total_users"],    "{:,.0f}"),
            ("Conversions",       ga4["total_conv"],     "{:.0f}"),
            ("Bounce Rate",       ga4["avg_bounce"],     "{:.1%}"),
            ("Avg Session (sec)", ga4["avg_dur"],        "{:.0f}s"),
        ]
        for label, val, fmt in g4kpis:
            write_kpi_row(ws, row, label, val, fmt)
            row += 1

    # ── Sheet 2: Issues ──
    ws2 = wb.create_sheet("Issues & Fixes")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 50
    header_style(ws2, 1, 4, "Issues & Prioritized Fixes", color)

    headers = ["Priority", "Issue", "Detail", "Recommended Fix"]
    for i, h in enumerate(headers, 1):
        c = ws2.cell(row=2, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="dddddd")

    all_issues = (gads.get("issues",[]) if "error" not in gads else []) + \
                 (ga4.get("issues",[])   if "error" not in ga4  else [])
    priority_colors = {"CRITICAL":"FF4136","HIGH":"FF6B35","MEDIUM":"FFD700","LOW":"90EE90"}

    for i, issue in enumerate(all_issues, 3):
        pc = priority_colors.get(issue["priority"], "FFFFFF")
        ws2.cell(row=i, column=1, value=issue["priority"]).fill = PatternFill("solid", fgColor=pc)
        ws2.cell(row=i, column=1).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=issue["issue"])
        ws2.cell(row=i, column=3, value=issue["detail"])
        ws2.cell(row=i, column=4, value=issue["fix"])
        ws2.row_dimensions[i].height = 22
        for col in range(1, 5):
            ws2.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="center")

    # ── Sheet 3: Campaigns ──
    if "error" not in gads and gads.get("campaigns"):
        ws3 = wb.create_sheet("Campaign Breakdown")
        headers3 = ["Campaign","Status","Strategy","Spend","Clicks","Impressions","Conv","CPC","CTR","CPA"]
        ws3.column_dimensions["A"].width = 35
        for i in range(2, 11):
            ws3.column_dimensions[get_column_letter(i)].width = 14
        header_style(ws3, 1, len(headers3), "Campaign Breakdown — Last 30 Days", color)
        for j, h in enumerate(headers3, 1):
            c = ws3.cell(row=2, column=j, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="dddddd")
        for i, camp in enumerate(gads["campaigns"], 3):
            vals = [
                camp["name"], camp["status"], camp["strategy"],
                f"${camp['spend']:.2f}", int(camp["clicks"]), int(camp["impressions"]),
                f"{camp['conversions']:.1f}", f"${camp['cpc']:.2f}",
                f"{camp['ctr']:.2%}", f"${camp['cpa']:.2f}"
            ]
            for j, v in enumerate(vals, 1):
                ws3.cell(row=i, column=j, value=v)

    fname = out_dir / f"{client_key}_audit_{TODAY}.xlsx"
    wb.save(fname)
    return fname


# ─── PDF BUILDER ─────────────────────────────────────────────────────────────
def build_pdf(client_key, client_cfg, gads, ga4, out_dir):
    fname = out_dir / f"{client_key}_audit_{TODAY}.pdf"
    PAGE_W, PAGE_H = A4
    L_MAR = R_MAR = 20*mm
    USABLE_W = PAGE_W - L_MAR - R_MAR

    doc = SimpleDocTemplate(
        str(fname),
        pagesize=A4,
        leftMargin=L_MAR, rightMargin=R_MAR,
        topMargin=20*mm, bottomMargin=20*mm
    )

    styles = getSampleStyleSheet()
    brand  = colors.HexColor(client_cfg["color"])
    accent = colors.HexColor(client_cfg["accent"])

    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=18, textColor=brand, spaceAfter=4)
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, textColor=brand, spaceAfter=3)
    BODY = ParagraphStyle("BODY", parent=styles["Normal"], fontSize=9, leading=13)
    SMALL = ParagraphStyle("SMALL", parent=styles["Normal"], fontSize=8, leading=11, textColor=colors.gray)

    story = []

    # Cover
    story.append(Paragraph(f"360° Daily Audit Report", H1))
    story.append(Paragraph(client_cfg["full_name"], ParagraphStyle("SUB", parent=styles["Normal"], fontSize=14, textColor=colors.gray)))
    story.append(Paragraph(f"Period: {DATE_FROM} → {DATE_TO} | Generated: {TODAY}", SMALL))
    story.append(HRFlowable(width=USABLE_W, color=brand, thickness=2, spaceAfter=8))

    def kpi_table(data, col_widths=None):
        cw = col_widths or [USABLE_W*0.45, USABLE_W*0.25, USABLE_W*0.30]
        t = Table(data, colWidths=cw, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), brand),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, accent]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.lightgrey),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ]))
        return t

    # Google Ads section
    story.append(Paragraph("Google Ads Performance", H2))
    if "error" in gads:
        story.append(Paragraph(f"⚠️ {gads['error']}", BODY))
    else:
        kpi_data = [
            ["Metric", "Value", "Benchmark"],
            ["Total Spend",   f"${gads['total_spend']:,.2f}", "—"],
            ["Total Clicks",  f"{gads['total_clicks']:,.0f}", "—"],
            ["Impressions",   f"{gads['total_impr']:,.0f}",  "—"],
            ["Conversions",   f"{gads['total_conv']:.1f}",   "—"],
            ["Avg CPC",       f"${gads['avg_cpc']:.2f}",     "$2–8"],
            ["Avg CTR",       f"{gads['avg_ctr']:.2%}",      "2–5%"],
            ["Avg CPA",       f"${gads['avg_cpa']:.2f}",     "Varies"],
            ["ROAS",          f"{gads['avg_roas']:.2f}x",    ">3x"],
        ]
        story.append(kpi_table(kpi_data))
        story.append(Spacer(1, 6))

        if gads.get("campaigns"):
            story.append(Paragraph("Campaign Breakdown", H2))
            camp_data = [["Campaign", "Spend", "Clicks", "Conv", "CTR", "CPA"]]
            for c in gads["campaigns"][:10]:
                camp_data.append([
                    c["name"][:35], f"${c['spend']:.0f}", f"{int(c['clicks'])}",
                    f"{c['conversions']:.1f}", f"{c['ctr']:.2%}", f"${c['cpa']:.0f}"
                ])
            cw = [USABLE_W*0.40, USABLE_W*0.12, USABLE_W*0.10, USABLE_W*0.10, USABLE_W*0.12, USABLE_W*0.16]
            story.append(kpi_table(camp_data, cw))

    story.append(Spacer(1, 8))

    # GA4 section
    story.append(Paragraph("GA4 Analytics", H2))
    if "error" in ga4:
        story.append(Paragraph(f"⚠️ {ga4['error']}", BODY))
    else:
        ga4_data = [
            ["Metric", "Value", "Benchmark"],
            ["Sessions",        f"{ga4['total_sessions']:,.0f}", "—"],
            ["Active Users",    f"{ga4['total_users']:,.0f}",    "—"],
            ["Conversions",     f"{ga4['total_conv']:.0f}",      "—"],
            ["Bounce Rate",     f"{ga4['avg_bounce']:.1%}",      "<60%"],
            ["Avg Session",     f"{ga4['avg_dur']:.0f}s",        ">90s"],
        ]
        story.append(kpi_table(ga4_data))

    story.append(Spacer(1, 10))

    # Issues & Fixes
    story.append(HRFlowable(width=USABLE_W, color=brand, thickness=1, spaceAfter=6))
    story.append(Paragraph("Issues & Prioritized Fixes", H2))

    all_issues = (gads.get("issues",[]) if "error" not in gads else []) + \
                 (ga4.get("issues",[])   if "error" not in ga4  else [])

    if not all_issues:
        story.append(Paragraph("✅ No critical issues detected. Continue monitoring.", BODY))
    else:
        issue_data = [["#", "Priority", "Issue", "Fix"]]
        p_colors = {"CRITICAL": colors.HexColor("#FF4136"), "HIGH": colors.HexColor("#FF6B35"),
                    "MEDIUM": colors.HexColor("#FFD700"),   "LOW":  colors.HexColor("#90EE90")}
        row_styles = []
        for idx, iss in enumerate(all_issues, 1):
            issue_data.append([str(idx), iss["priority"], iss["issue"][:40], iss["fix"][:60]])
            row_styles.append(("BACKGROUND", (1, idx), (1, idx), p_colors.get(iss["priority"], colors.white)))

        cw2 = [USABLE_W*0.05, USABLE_W*0.12, USABLE_W*0.35, USABLE_W*0.48]
        t2 = Table(issue_data, colWidths=cw2, repeatRows=1)
        base_style = [
            ("BACKGROUND",   (0,0), (-1,0), brand),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 7.5),
            ("GRID",         (0,0), (-1,-1), 0.3, colors.lightgrey),
            ("TOPPADDING",   (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
            ("WORDWRAP",     (0,0), (-1,-1), True),
        ] + row_styles
        t2.setStyle(TableStyle(base_style))
        story.append(t2)

    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Generated by DigiMinds Daily Audit System | {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", SMALL))

    doc.build(story)
    return fname


# ─── HTML DASHBOARD ──────────────────────────────────────────────────────────
def build_html(client_key, client_cfg, gads, ga4, out_dir):
    color  = client_cfg["color"]
    accent = client_cfg["accent"]
    name   = client_cfg["full_name"]

    all_issues = (gads.get("issues",[]) if "error" not in gads else []) + \
                 (ga4.get("issues",[])   if "error" not in ga4  else [])
    priority_badge = {
        "CRITICAL": '<span style="background:#FF4136;color:white;padding:2px 8px;border-radius:3px;font-size:11px;font-weight:bold;">CRITICAL</span>',
        "HIGH":     '<span style="background:#FF6B35;color:white;padding:2px 8px;border-radius:3px;font-size:11px;font-weight:bold;">HIGH</span>',
        "MEDIUM":   '<span style="background:#FFD700;color:#333;padding:2px 8px;border-radius:3px;font-size:11px;font-weight:bold;">MEDIUM</span>',
        "LOW":      '<span style="background:#90EE90;color:#333;padding:2px 8px;border-radius:3px;font-size:11px;font-weight:bold;">LOW</span>',
    }

    def kpi_card(label, val, sub=""):
        return f"""
        <div style="background:white;border-radius:8px;padding:16px 20px;box-shadow:0 2px 8px rgba(0,0,0,.08);min-width:140px;flex:1">
            <div style="font-size:11px;color:#888;text-transform:uppercase;letter-spacing:.5px">{label}</div>
            <div style="font-size:26px;font-weight:700;color:{color};margin:4px 0">{val}</div>
            <div style="font-size:11px;color:#aaa">{sub}</div>
        </div>"""

    gads_cards = ""
    if "error" not in gads:
        gads_cards = "".join([
            kpi_card("Spend",       f"${gads['total_spend']:,.2f}"),
            kpi_card("Clicks",      f"{gads['total_clicks']:,.0f}"),
            kpi_card("Conversions", f"{gads['total_conv']:.1f}"),
            kpi_card("CPC",         f"${gads['avg_cpc']:.2f}",  "avg"),
            kpi_card("CTR",         f"{gads['avg_ctr']:.2%}",   "avg"),
            kpi_card("CPA",         f"${gads['avg_cpa']:.2f}",  "avg"),
            kpi_card("ROAS",        f"{gads['avg_roas']:.2f}x"),
        ])
    else:
        gads_cards = f'<div style="color:#888;padding:20px">{gads["error"]}</div>'

    ga4_cards = ""
    if "error" not in ga4:
        ga4_cards = "".join([
            kpi_card("Sessions",   f"{ga4['total_sessions']:,.0f}"),
            kpi_card("Users",      f"{ga4['total_users']:,.0f}"),
            kpi_card("Conv",       f"{ga4['total_conv']:.0f}"),
            kpi_card("Bounce",     f"{ga4['avg_bounce']:.1%}",  "<60% good"),
            kpi_card("Avg Sess",   f"{ga4['avg_dur']:.0f}s",    ">90s good"),
        ])
    else:
        ga4_cards = f'<div style="color:#888;padding:20px">{ga4["error"]}</div>'

    issues_html = ""
    for iss in all_issues:
        issues_html += f"""
        <tr>
            <td style="padding:10px 12px">{priority_badge.get(iss['priority'],'')}</td>
            <td style="padding:10px 12px;font-weight:600">{iss['issue']}</td>
            <td style="padding:10px 12px;color:#555;font-size:13px">{iss['detail']}</td>
            <td style="padding:10px 12px;color:#1a4f8a;font-size:13px">💡 {iss['fix']}</td>
        </tr>"""
    if not issues_html:
        issues_html = '<tr><td colspan="4" style="padding:20px;color:green;text-align:center">✅ No critical issues detected</td></tr>'

    camp_rows = ""
    if "error" not in gads and gads.get("campaigns"):
        for c in gads["campaigns"][:12]:
            flag = "🔴" if c["ctr"] < 0.01 else ("🟡" if c["ctr"] < 0.02 else "🟢")
            camp_rows += f"""
            <tr>
                <td style="padding:8px 12px">{c['name'][:40]}</td>
                <td style="padding:8px 12px">{c['status']}</td>
                <td style="padding:8px 12px">${c['spend']:.2f}</td>
                <td style="padding:8px 12px">{int(c['clicks'])}</td>
                <td style="padding:8px 12px">{c['conversions']:.1f}</td>
                <td style="padding:8px 12px">{flag} {c['ctr']:.2%}</td>
                <td style="padding:8px 12px">${c['cpa']:.2f}</td>
            </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} — Daily Audit {TODAY}</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0 }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; background:#f4f6f9; color:#222 }}
  .header {{ background:{color}; color:white; padding:24px 32px }}
  .header h1 {{ font-size:22px; font-weight:700 }}
  .header p {{ font-size:13px; opacity:.8; margin-top:4px }}
  .section {{ margin:20px 32px }}
  .section-title {{ font-size:15px; font-weight:700; color:{color}; margin-bottom:12px; padding-bottom:4px; border-bottom:2px solid {color} }}
  .kpi-row {{ display:flex; gap:12px; flex-wrap:wrap; margin-bottom:8px }}
  table {{ width:100%; border-collapse:collapse; background:white; border-radius:8px; overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,.06) }}
  thead tr {{ background:{color}; color:white }}
  th {{ padding:10px 12px; text-align:left; font-size:12px; font-weight:600 }}
  tbody tr:nth-child(even) {{ background:{accent} }}
  tbody tr:hover {{ background:#e8f4fd }}
  .footer {{ text-align:center; color:#aaa; font-size:11px; padding:24px }}
</style>
</head>
<body>
<div class="header">
  <h1>📊 {name} — Daily PPC Audit</h1>
  <p>Period: {DATE_FROM} → {DATE_TO} | Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
</div>

<div class="section">
  <div class="section-title">🎯 Google Ads KPIs (Last 30 Days)</div>
  <div class="kpi-row">{gads_cards}</div>
</div>

<div class="section">
  <div class="section-title">📈 GA4 Analytics (Last 30 Days)</div>
  <div class="kpi-row">{ga4_cards}</div>
</div>

<div class="section">
  <div class="section-title">⚠️ Issues & Prioritized Fixes ({len(all_issues)} found)</div>
  <table>
    <thead><tr><th>Priority</th><th>Issue</th><th>Detail</th><th>Fix</th></tr></thead>
    <tbody>{issues_html}</tbody>
  </table>
</div>

{'<div class="section"><div class="section-title">📋 Campaign Breakdown</div><table><thead><tr><th>Campaign</th><th>Status</th><th>Spend</th><th>Clicks</th><th>Conv</th><th>CTR</th><th>CPA</th></tr></thead><tbody>' + camp_rows + '</tbody></table></div>' if camp_rows else ''}

<div class="footer">DigiMinds Daily Audit System | {TODAY}</div>
</body></html>"""

    fname = out_dir / f"{client_key}_dashboard_{TODAY}.html"
    fname.write_text(html, encoding="utf-8")
    return fname


# ─── MAIN RUNNER ─────────────────────────────────────────────────────────────
def run_client(client_key, client_cfg):
    now      = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H-%M-%S")

    out_dir = DOWNLOADS / client_cfg["full_name"].replace(" ", "_") / date_str / time_str
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n{'='*60}")
    print(f"Client: {client_cfg['full_name']}")
    print(f"Output: {out_dir}")

    # Pull data
    print("  Pulling Google Ads…")
    gads_rows = windsor_pull("google_ads", GADS_FIELDS,
                             client_cfg["gads_accounts"], DATE_FROM, DATE_TO)
    print(f"  → {len(gads_rows)} rows")

    print("  Pulling GA4…")
    ga4_rows  = windsor_pull("googleanalytics4", GA4_FIELDS,
                             client_cfg["ga4_accounts"], DATE_FROM, DATE_TO)
    print(f"  → {len(ga4_rows)} rows")

    # Inject client-specific benchmarks as metadata on first row
    benchmarks = {
        "Ortho":     {"cpa_target": 80,  "ctr_target": 0.05, "is_target": 0.40},
        "HeatWeave": {"cpa_target": 120, "ctr_target": 0.03, "is_target": 0.30},
        "Tack":      {"cpa_target": 60,  "ctr_target": 0.04, "is_target": 0.35},
    }
    if gads_rows and client_key in benchmarks:
        gads_rows[0]["__client_benchmarks__"] = benchmarks[client_key]

    gads = analyze_gads(gads_rows)
    ga4  = analyze_ga4(ga4_rows)

    print(f"  Issues: {len(gads.get('issues', []))+len(ga4.get('issues', []))}")

    # Build reports
    print("  Building XLSX…")
    xlsx = build_xlsx(client_key, client_cfg, gads, ga4, out_dir)
    print(f"  → {xlsx.name}")

    print("  Building PDF…")
    pdf = build_pdf(client_key, client_cfg, gads, ga4, out_dir)
    print(f"  → {pdf.name}")

    print("  Building HTML…")
    html = build_html(client_key, client_cfg, gads, ga4, out_dir)
    print(f"  → {html.name}")

    return {"xlsx": xlsx, "pdf": pdf, "html": html, "issues": len(gads.get("issues",[])+ga4.get("issues",[]))}


def main():
    print(f"DigiMinds Daily Ads Audit — {TODAY}")
    print(f"Windsor API: {'✓ configured' if WINDSOR_API_KEY != 'YOUR_WINDSOR_API_KEY_HERE' else '✗ MISSING — set WINDSOR_API_KEY env var'}")

    results = {}
    for key, cfg in CLIENTS.items():
        try:
            results[key] = run_client(key, cfg)
        except Exception as e:
            print(f"  ERROR on {key}: {e}")
            results[key] = {"error": str(e)}

    print(f"\n{'='*60}")
    print("AUDIT COMPLETE")
    for key, r in results.items():
        if "error" in r:
            print(f"  ✗ {key}: {r['error']}")
        else:
            print(f"  ✓ {key}: {r['issues']} issues | {r['pdf'].parent}")


if __name__ == "__main__":
    main()
